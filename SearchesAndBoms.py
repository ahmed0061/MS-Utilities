import openpyxl
import re





# get file objects
file = r'D:\Desktop\py\SearchesAndBoms.xlsx'
sheet = "Sheet1"

f_obj = openpyxl.load_workbook(file)
f_sheet_obj = f_obj.get_sheet_by_name(sheet)
rows = f_sheet_obj.get_highest_row()        # include header



################################################  FIRST COLUMN
# map for undesired data may found in first_col .
caption_map = {
" Authorized Distributor":"",
" ECIA (NEDA) Member •":"",
" Manufacturer Direct – Inventory Available for Immediate and Future Delivery":"",
" Manufacturer Direct - Purchase at the lowest online price* on TI.com":"",
" ECIA (NEDA)":"",
" Manufacturer Direct • Free Shipping":"",
" Free 24 Hour Samples":"",
}


# replace multiple substitutions in string .. stackoverflow
def multiple_replace(dict, text):
  # Create a regular expression  from the dictionary keys
  regex = re.compile("(%s)" % "|".join(map(re.escape, dict.keys())))

  # For each match, look-up corresponding value in dictionary
  return regex.sub(lambda mo: dict[mo.string[mo.start():mo.end()]], text) 




# first column generator
first_col = ( f_sheet_obj.cell(row=i, column= 1).value  for i in range(1, rows + 1) )
# cleanining and list data
first_col = [ multiple_replace(caption_map, cell ) for cell in  first_col  ]


# save first column
for i in range(1, len(first_col)+1):
    f_sheet_obj.cell(row=i, column=10).value = first_col[i-1]

# name header
f_sheet_obj.cell(row=1, column=10).value = "Company"

f_obj.save("test_regex.xlsx")



# delete undesired data to free some ram .
del caption_map, first_col



################################################ SECOND COLUMN
# read second col. cells
sec_column = [f_sheet_obj.cell(row=i, column= 2).value for i in range(1, rows + 1) ]


# parts are first value in splitting sec_column
parts_col = [ val[0] if val else "" for val in ( cell.split(' DISTI # ') for cell in sec_column ) ]

# sku are second value in splitting sec_column 
sku_col = [ val[1] if len(val)==2 else "" for val in ( cell.split(' DISTI # ') for cell in sec_column ) ]


# save parts and sku
for i in range(1, len(parts_col)+1):
    f_sheet_obj.cell(row=i, column=11).value = parts_col[i-1]
    f_sheet_obj.cell(row=i, column=12).value = sku_col[i-1]

# name headers
f_sheet_obj.cell(row=1, column=11).value = "PartNumber"
f_sheet_obj.cell(row=1, column=12).value = "SKU"


f_obj.save("test_regex.xlsx")


# delete undesired data to free some ram .
del sec_column, parts_col, sku_col



################################################ THIRD COLUMN
# name header
Data_Manufacturer = [f_sheet_obj.cell(row=i, column= 3).value for i in range(1, rows + 1) ]

for i in range(1, len(Data_Manufacturer)+1):
    f_sheet_obj.cell(row=i, column=13).value = Data_Manufacturer[i-1]


f_sheet_obj.cell(row=1, column=13).value = "Data.Manufacturer"
f_obj.save("test_regex.xlsx")


del Data_Manufacturer

################################################ Fourth COLUMN
# name header
Data_Description = [f_sheet_obj.cell(row=i, column= 4).value for i in range(1, rows + 1) ]

for i in range(1, len(Data_Description)+1):
    f_sheet_obj.cell(row=i, column=14).value = Data_Description[i-1]


f_sheet_obj.cell(row=1, column=14).value = "Data.Manufacturer"
f_obj.save("test_regex.xlsx")


del Data_Description

################################################ Fifth COLUMN

# read fifth col.
fif_col = ( f_sheet_obj.cell(row=i, column= 5).value for i in range(1, rows + 1) )

# filter only numbers for to get stock  .

# stock_col = [ re.compile(r'[0-9]*').search(text).group() if text else "" for text in fif_col ]
# stock_col = [ str(  re.compile(r'\d+').findall(text)  ) if text else "" for text in fif_col ]
stock_col = [ re.compile(r'[0-9]*(,*)[0-9]*|\d*').search(text).group() if text else "" for text in fif_col ]   ## working ok except for case contain dash in between because search method catch 1st occurance .

# save stock and sku
for i in range(1, len(stock_col)+1):
    f_sheet_obj.cell(row=i, column=15).value = stock_col[i-1]

f_sheet_obj.cell(row=1, column=15).value = "Stock"

f_obj.save("test_regex.xlsx")



del fif_col, stock_col



################################################ Sixth COLUMN
six_col = [ f_sheet_obj.cell(row=i, column= 6).value for i in range(1, rows + 1) ]
six_col[0] = "PriceBreak$Price"
# print(list(six_col))

# six_col_splitted =  ( cell.split("$") if cell else "" for cell in six_col )

PriceBreak = [  val[0] for val in ( cell.split("$") if cell else "" for cell in six_col ) ]
Price = [  val[1] for val in ( cell.split("$") if cell else "" for cell in six_col ) ]

# save PriceBreak, Price
for i in range(1, len(PriceBreak)+1):
    f_sheet_obj.cell(row=i, column=16).value = PriceBreak[i-1]
    f_sheet_obj.cell(row=i, column=17).value = Price[i-1]
f_obj.save("test_regex.xlsx")


del six_col, PriceBreak, Price


print("="*10 + " Done "+"="*10)


























