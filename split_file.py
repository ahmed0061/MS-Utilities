# divide large excel file into small files
# failed because memory error




import openpyxl



# split file into small files

# file = "test_split.xlsx"
file = "stocked-imp.xlsx"
file_obj = openpyxl.load_workbook(file)
file_sheet = file_obj.get_sheet_by_name("Sheet1")
file_rows = file_sheet.get_highest_row()



columns_names = []
i = 1
flag = True
while flag:
    column_name = file_sheet.cell(row = 1, column = i).value
    if not column_name:
        flag = False
    else:
        columns_names.append(column_name)
        i += 1




# # create list for each header name
columns = { h:[] for h in columns_names   }



# loop through column names
for col in columns_names:
    columns[col] = [  file_sheet.cell(row=j, column=columns_names.index(col)+1 ).value  for j in range(1, file_rows+1 ) ]



for k, v in columns.items():
    print(k)
    print(v)