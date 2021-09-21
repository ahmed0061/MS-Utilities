"""
MODEULE FUNCTION:

fills blank cells in excel file with max value in column before,
cells are in predetermined columns that we will loop through them

MODEULE STATUS:
Working
"""

import openpyxl
import pandas as pd  
import os


# vars
file = "test.xlsx"
sheet = "Sheet1"



# get file data
file_obj = openpyxl.load_workbook(file)
sheet_obj = file_obj.get_sheet_by_name(sheet)
rows = sheet_obj.get_highest_row()


# get max value from last column
def get_last_max(rows, sheet_obj, last_column_index):
	# create last column list
	last_column = []
	# loop through last column cells
	for i in range(2, rows+1):
		# check if cell has not int
		if not isinstance(last_column.append(sheet_obj.cell(row = i, column = last_column_index ).value), int):
			last_column.append(0)
		else:
			last_column.append(sheet_obj.cell(row = i, column = last_column_index ).value)
	max_val = max(last_column)
	return max_val

# create indication for columns that has blank cells
blank_columns = [i for i in range(5, 11)]

# loop through each column in blank_columns
for col in blank_columns:
	max_last_value = get_last_max(rows, sheet_obj, last_column_index=col-1)
	for i in range(2, rows+1):
		# if blank cell found
		if sheet_obj.cell(row = i, column = col ).value == None:
			sheet_obj.cell(row = i, column = col ).value = max_last_value

file_obj.save(file)






