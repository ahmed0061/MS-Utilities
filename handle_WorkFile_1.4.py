import openpyxl

# Standard variables .
path = r'C:\Users\ahmed.mosaad\Desktop\Rs\\ExcelUtils\\file_1.0\\'
sheet='Sheet1'
flag_name='IMPORTER'
f1='dm'
f2='nm'


def open_file(file_name, sheet='Sheet1'):
	"""Return file object, sheet object and rows"""
	file = path + file_name +'.xlsx'
	file_obj = openpyxl.load_workbook(file)
	file_sheet = file_obj.get_sheet_by_name(sheet)
	file_rows = file_sheet.get_highest_row()
	return file, file_obj, file_sheet, file_rows

def get_headers(file_sheet, sheet='Sheet1'):
	"""Return list names for columns""" 
	headers = []
	i = 1
	flag = True
	while flag:
		column_name = file_sheet.cell(row = 1, column = i).value
		if not column_name:
			flag = False
		else:
			headers.append(column_name)
			i += 1
	return headers

def get_flag_data(file_sheet, file_rows, headers, flag_name='IMPORTER'):
	"""returns flag column list"""
	flag_data = []
	for i in range(1, file_rows+1):
		flag_data.append(file_sheet.cell(row=i, column= headers.index(flag_name)+1).value)  # list index begins 0
	# delete item "IMPORTER" from  target_column_data
	del flag_data[0]
	return flag_data

def get_flag_indecies(flag_data, f1='dm', f2='nm'):
	# indecies for options, i.e get row number for existence for each option
	f1_indecies = []
	f2_indecies = []
	flag_dict = {}								# this dictionary: to get index each option has ,so can know indecies for rows where option exist .
	for i in range(len(flag_data)):
		flag_dict[i] = flag_data[i]
	for k,v in flag_dict.items():
		if v == f1:
			f1_indecies.append(k)
		elif v == f2:
			f2_indecies.append(k)
	# shifting indecies by 2 :as list index starts at 0, and excel index starts at 2
	f1_indecies = [i+2 for i in f1_indecies]
	f2_indecies = [i+2 for i in f2_indecies]
	return f1_indecies, f2_indecies

def clean_file(file, file_obj, file_sheet, file_rows):
	# delete last data from file
	file_indecies = [i for i in range(1, 40)]
	for j in range(2, file_rows+1):
		for file_index in file_indecies:
			file_sheet.cell(row = j, column = file_index ).value = None
	file_obj.save(file)
	print('\nfile  cleaned')

def get_mask_data(f1_indecies, file_sheet, headers):
	ZPartNumber, ZCompanyName, ZReplacementPart, ZReplacementCompany, obsCode, obsReas, hasRepl, OnlineSource, ZURLSource= ([] for i in range(9))
	for f in f1_indecies:
		ZPartNumber.append(file_sheet.cell(row=f, column=headers.index('ZPartNumber')+1).value)
		ZCompanyName.append(file_sheet.cell(row=f, column=headers.index('ZCompanyName')+1).value)
		ZReplacementPart.append(file_sheet.cell(row=f, column=headers.index('ZReplacementPart')+1).value)
		obsCode.append(file_sheet.cell(row=f, column=headers.index('obsCode')+1).value)
		obsReas.append(file_sheet.cell(row=f, column=headers.index('obsReas')+1).value)
		hasRepl.append(file_sheet.cell(row=f, column=headers.index('hasRepl')+1).value)
		OnlineSource.append(file_sheet.cell(row=f, column=headers.index('OnlineSource')+1).value)
		ZURLSource.append(file_sheet.cell(row=f, column=headers.index('ZURLSource')+1).value)
	ZReplacementCompany = ZCompanyName
	print('\nCopying Mask data Done..'+str(len(ZPartNumber))+' row\n')
	return ZPartNumber, ZCompanyName, ZReplacementPart, ZReplacementCompany, obsCode, obsReas, hasRepl, OnlineSource, ZURLSource

def get_ZReplacementType_by_code(obsCode):
	*_ , code_sheet, code_rows = open_file('codes')
	code_list = []
	replacement = []
	codes_dict = {}
	ZReplacementType = []
	for z in range(2, code_rows+1):
		code_list.append(code_sheet.cell(row=z, column=1).value)
		replacement.append(code_sheet.cell(row=z, column=2).value)
	for k, v in zip(code_list, replacement):
		codes_dict[k] = v
	for t in obsCode:
		ZReplacementType.append(codes_dict[t]) 
	return ZReplacementType

def make_mask_file(*args):
	for i in range(2, len(ZPartNumber)+2):
		mask_sheet.cell(row=i, column=1).value = ZPartNumber[i-2]
		mask_sheet.cell(row=i, column=2).value = ZCompanyName[i-2]
		mask_sheet.cell(row=i, column=3).value = ZReplacementPart[i-2]
		mask_sheet.cell(row=i, column=4).value = ZReplacementCompany[i-2]
		mask_sheet.cell(row=i, column=5).value = ZReplacementType[i-2]
		mask_sheet.cell(row=i, column=7).value = 'Exact'
		mask_sheet.cell(row=i, column=8).value = 'OutOfZ2'
		mask_sheet.cell(row=i, column=9).value = OnlineSource[i-2]
		mask_sheet.cell(row=i, column=10).value = ZURLSource[i-2]
		mask_sheet.cell(row=i, column=11).value = 'Supplier Source'
	mask_obj.save(mask)
	print('\nMask file Done ..'+str(len(ZPartNumber))+'row')

def get_notmatch_data(f2_indecies, file_sheet, headers):
	ZPartNumber_nm, ZCompanyName_nm, ZReplacementPart_nm, ZReplacementCompany_nm, OnlineSource_nm, ZURLSource_nm  = ([] for i in range(6))
	for i in f2_indecies:
		ZPartNumber_nm.append(file_sheet.cell(row=i, column=headers.index('ZPartNumber')+1).value)
		ZCompanyName_nm.append(file_sheet.cell(row=i, column=headers.index('ZCompanyName')+1).value)
		ZReplacementPart_nm.append(file_sheet.cell(row=i, column=headers.index('ZReplacementPart')+1).value)
		OnlineSource_nm.append(file_sheet.cell(row=i, column=headers.index('OnlineSource')+1).value)
		ZURLSource_nm.append(file_sheet.cell(row=i, column=headers.index('ZURLSource')+1).value)
	ZReplacementCompany_nm = ZCompanyName_nm
	
	comment_nm = []
	for part in ZReplacementPart_nm:
		if part  == 'N/A':
			comment_nm.append('')
		else :
			comment_nm.append('Part')

	print('\nCopying NotMatch data Done..' + str(len(ZPartNumber_nm)) + ' row')
	return ZPartNumber_nm, ZCompanyName_nm, ZReplacementPart_nm, ZReplacementCompany_nm, OnlineSource_nm, ZURLSource_nm, comment_nm

def make_notmatch_file(*args):
	for i in range(2, len(ZPartNumber_nm)+2):
		NotMatch_sheet.cell(row=i, column=1).value = ZPartNumber_nm[i-2]
		NotMatch_sheet.cell(row=i, column=2).value = ZCompanyName_nm[i-2]
		NotMatch_sheet.cell(row=i, column=3).value = ZReplacementPart_nm[i-2]
		NotMatch_sheet.cell(row=i, column=4).value = ZReplacementCompany_nm[i-2]
		NotMatch_sheet.cell(row=i, column=5).value = comment_nm[i-2]
		NotMatch_sheet.cell(row=i, column=6).value = OnlineSource_nm[i-2]
		NotMatch_sheet.cell(row=i, column=7).value = ZURLSource_nm[i-2]
	NotMatch_obj.save(NotMatch)
	print('\nNotMatch file Done ..'+str(len(ZPartNumber_nm))+'row')

def get_obs_data(file_sheet, file_rows,  headers):
	ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs= ([] for i in range(10))
	for i in range(2, file_rows+1):
		ZPartNumber_obs.append(file_sheet.cell(row=i, column=headers.index('ZPartNumber')+1).value)
		ZCompanyName_obs.append(file_sheet.cell(row=i, column=headers.index('ZCompanyName')+1).value)
		LCStatus_obs.append(file_sheet.cell(row=i, column=headers.index('ZLC')+1).value)
		ObsolescenceCode_obs.append(file_sheet.cell(row=i, column=headers.index('obsCode')+1).value)
		ObsolescenceReason_obs.append(file_sheet.cell(row=i, column=headers.index('obsReas')+1).value)
		HasReplacement_obs.append(file_sheet.cell(row=i, column=headers.index('hasRepl')+1).value)	
		OnLineSource1_obs.append(file_sheet.cell(row=i, column=headers.index('OnlineSource')+1).value)
		OffLineSource1_obs.append(file_sheet.cell(row=i, column=headers.index('ZURLSource')+1).value)
		SourceType_obs.append(file_sheet.cell(row=i, column=headers.index('SourceType')+1).value)

	for code, hasReplce in zip(ObsolescenceCode_obs, HasReplacement_obs):
		if code == 'OR20' or code =='OR7' or code =='OR11' or code =='OR29' or hasReplce =='No':
			Critical_obs.append('Critical')
		else:
			Critical_obs.append('Non Critical')
	print('\nCopying obs data Done..' + str(len(ZPartNumber_obs)) + ' row')
	return ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs

def make_obs_file(*args):
	for i in range(2, len(ZPartNumber_obs)+2):
		obs_sheet.cell(row=i, column=1).value = ZPartNumber_obs[i-2]
		obs_sheet.cell(row=i, column=2).value = ZCompanyName_obs[i-2]
		obs_sheet.cell(row=i, column=3).value = LCStatus_obs[i-2]
		obs_sheet.cell(row=i, column=4).value = ObsolescenceCode_obs[i-2]
		obs_sheet.cell(row=i, column=5).value = ObsolescenceReason_obs[i-2]
		obs_sheet.cell(row=i, column=6).value = HasReplacement_obs[i-2]
		obs_sheet.cell(row=i, column=7).value = Critical_obs[i-2]
		obs_sheet.cell(row=i, column=8).value = OnLineSource1_obs[i-2]
		obs_sheet.cell(row=i, column=9).value = OffLineSource1_obs[i-2]
		obs_sheet.cell(row=i, column=12).value = SourceType_obs[i-2]
	obs_obj.save(obs)
	print('\nobs file Done ..'+str(len(ZPartNumber_obs)) + ' row\n')

def get_partsRepInsert_data(file_sheet, file_rows,  headers):
	ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs = get_obs_data(file_sheet, file_rows,  headers)
	SourceType = []
	for url, hasRep, code in zip(OffLineSource1_obs, HasReplacement_obs, ObsolescenceCode_obs):
		if not url and code != 'OR20':
			SourceType.append('Z-Source')
		elif not url :
			SourceType.append('')
		elif hasRep=='NotMatch':
			SourceType.append('NotMatchedSource')
		else:
			SourceType.append('Supplier Source')
	print('\nCopying partsRepInsert data Done..' + str(len(ZPartNumber_obs)) + ' row')
	return ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs, SourceType

def get_NewReplaceType(obsCode):
	*_ , NewRep_sheet, NewRep_rows = open_file('codes_new')
	#to store data from file
	code = []
	Newreplace = []
	ReplaceFeature = []
	codetoNewreplace_dict = {}
	codetoReplaceFeature_dict = {}
	# to return for file .
	NewReplacementType = []
	ReplacementFeature = []

	# read data from file and store in lists .
	for z in range(2, NewRep_rows+1):
		code.append(NewRep_sheet.cell(row=z, column=1).value)
		Newreplace.append(NewRep_sheet.cell(row=z, column=4).value)
		ReplaceFeature.append(NewRep_sheet.cell(row=z, column=5).value)
	#make dict where code is always key
	for k, v in zip(code, Newreplace):
		codetoNewreplace_dict[k] = v
	for m, n in zip(code, ReplaceFeature):
		codetoReplaceFeature_dict[m] = n
	# get required data according to obsCode list .
	for t in obsCode:
		NewReplacementType.append(codetoNewreplace_dict[t]) 
		ReplacementFeature.append(codetoReplaceFeature_dict[t])

	return NewReplacementType, ReplacementFeature

def handle_exportedMask():
	f, f_obj, f_sheet, f_row = open_file('ex_dm')
	part, reptype, conca_partAndtype, ZReplacementPart, ZReplacementCompany, Diff_Features, PinToPin = ([] for a in range(7)) 
	for b in range(2, f_row+1):
		part.append(f_sheet.cell(row=b, column=1).value)
		reptype.append(f_sheet.cell(row=b, column=7).value)
		ZReplacementPart.append(f_sheet.cell(row=b, column=4).value)
		ZReplacementCompany.append(f_sheet.cell(row=b, column=5).value)
		Diff_Features.append(f_sheet.cell(row=b, column=12).value)
		PinToPin.append(f_sheet.cell(row=b, column=13).value)

	conca_partAndtype = [c+d for c,d in zip(part, reptype)]
	return conca_partAndtype, ZReplacementPart, ZReplacementCompany, Diff_Features, PinToPin

def handle_exportedNm():
	f, f_obj, f_sheet, f_row = open_file('ex_nm')
	partANDrep_nm = {}
	for b in range(2, f_row+1):
		if f_sheet.cell(row=b, column=4).value != 'N/A':
			partANDrep_nm[f_sheet.cell(row=b, column=1).value] = f_sheet.cell(row=b, column=4).value
	return partANDrep_nm


def handle_exportedfilesANDpartsRepInsert_file(ObsolescenceCode_obs, ZPartNumber_obs, HasReplacement_obs, ZCompanyName_obs):
	conca_exported, ZReplacementPart, ZReplacementCompany, Diff_Features, PinToPin = handle_exportedMask()
	partANDrep_nm = handle_exportedNm()
	conca_exportedANDrepl = {a:b for a, b in zip(conca_exported, ZReplacementPart)}
	conca_exportedANDreplcomp = {c:d for c, d in zip(conca_exported, ZReplacementCompany)}
	conca_exportedANDdiff = {e:f for e, f in zip(conca_exported, Diff_Features)}
	conca_exportedANDpin = {g:h for g, h in zip(conca_exported, PinToPin)}

	reptype = get_ZReplacementType_by_code(ObsolescenceCode_obs)
	conca = [a+b for a,b in zip(ZPartNumber_obs, reptype)]

	ZReplacementPart, ZReplacementCompany, DiffFeatures, PinToPin = ([] for a in range(4))

	for con in conca:
		if con in conca_exported:
			ZReplacementPart.append(conca_exportedANDrepl[con])
			ZReplacementCompany.append(conca_exportedANDreplcomp[con])
			DiffFeatures.append(conca_exportedANDdiff[con])
			PinToPin.append(conca_exportedANDpin[con])
		else:
			ZReplacementPart.append('')
			ZReplacementCompany.append('')
			DiffFeatures.append('')
			PinToPin.append('')


	for a in range(0, len(HasReplacement_obs)) :
		if HasReplacement_obs[a] == 'NotMatch':
			ZReplacementPart[a] = partANDrep_nm[ZPartNumber_obs[a]]
			ZReplacementCompany[a] = ZCompanyName_obs[a]
	return ZReplacementPart, ZReplacementCompany, DiffFeatures, PinToPin


def get_Exceptionflag(ObsolescenceCode_obs, OffLineSource1_obs):
	# countif algorithm in excel .
	urlANDobs_distinct = []
	for url, obs in zip(OffLineSource1_obs, ObsolescenceCode_obs):
		urlANDobs_distinct.append(url+'$'+obs)
	urlANDobs_distinct = set(urlANDobs_distinct)
	urlANDobs_distinct_1 = []

	for item in urlANDobs_distinct:
		urlANDobs_distinct_1.append(item.split('$'))

	urlANDobs_distinct_2 = []
	for list_ in urlANDobs_distinct_1:
		urlANDobs_distinct_2.append(list_[0])

	urlTocounts_dict = {}
	for url in urlANDobs_distinct_2:
		urlTocounts_dict[url] = urlANDobs_distinct_2.count(url)

	urlTocounts_dict_1 = {}
	for k, v in urlTocounts_dict.items():
		if v > 1:
			urlTocounts_dict_1[k] = "1"
		else:
			urlTocounts_dict_1[k] = "0"

	Exceptionflag = []
	for OffLine in OffLineSource1_obs:
		Exceptionflag.append(urlTocounts_dict_1[OffLine])
	return Exceptionflag

def make_partsRepInsert_file(*args, exportedMask=False):
	for i in range(2, len(ZPartNumber_obs)+2):
		partsRepInsert_sheet.cell(row=i, column=1).value = ZPartNumber_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=2).value = ZCompanyName_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=5).value = ObsolescenceCode_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=6).value = ObsolescenceReason_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=8).value = NewReplacementType[i-2]
		partsRepInsert_sheet.cell(row=i, column=9).value = ReplacementFeature[i-2]
		# partsRepInsert_sheet.cell(row=i, column=10).value = Display on Portal[i-2]
		partsRepInsert_sheet.cell(row=i, column=11).value = SourceType[i-2]
		partsRepInsert_sheet.cell(row=i, column=12).value = OnLineSource1_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=13).value = OffLineSource1_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=16).value = OnLineSource1_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=17).value = OffLineSource1_obs[i-2]
		partsRepInsert_sheet.cell(row=i, column=20).value = '1'
		partsRepInsert_sheet.cell(row=i, column=21).value = '1'
		# partsRepInsert_sheet.cell(row=i, column=22).value = Exceptionflag[i-2]
		# partsRepInsert_sheet.cell(row=i, column=23).value = PinToPin[i-2]						
		partsRepInsert_sheet.cell(row=i, column=25).value = HasReplacement_obs[i-2]

		if exportedMask:
			partsRepInsert_sheet.cell(row=i, column=3).value = ZReplacementPart[i-2]
			partsRepInsert_sheet.cell(row=i, column=4).value = ZReplacementCompany[i-2]
			partsRepInsert_sheet.cell(row=i, column=7).value = DiffFeatures[i-2]
			partsRepInsert_sheet.cell(row=i, column=23).value = PinToPin[i-2]
	partsRepInsert_obj.save(partsRepInsert)
	print('\npartsRepInsert file Done ..'+str(len(ZPartNumber_obs)) + ' row\n')





###########################################################################################################################################################################################
																				# WORK CYCLE #
###########################################################################################################################################################################################
# work file
work, work_obj, work_sheet, work_rows = open_file('Work')
headers = get_headers(work_sheet)
flag_data = get_flag_data(work_sheet, work_rows, headers)
f1_indecies, f2_indecies = get_flag_indecies(flag_data)


##clean files
files = ['obs', 'Mask', 'NotMatch', 'PartsReplacementInsert']
for f in files:
	file, file_obj, file_sheet, file_rows = open_file(f)
	clean_file(file, file_obj, file_sheet, file_rows)	

# mask file
mask, mask_obj, mask_sheet, mask_rows = open_file('Mask')
ZPartNumber, ZCompanyName, ZReplacementPart, ZReplacementCompany, obsCode, obsReas, hasRepl, OnlineSource, ZURLSource = get_mask_data(f1_indecies, work_sheet, headers)
ZReplacementType = get_ZReplacementType_by_code(obsCode)
make_mask_file(mask_sheet, ZPartNumber, ZCompanyName, ZReplacementPart, ZReplacementCompany, obsCode, obsReas, hasRepl, OnlineSource, ZURLSource, ZReplacementType)


# notmatch file
NotMatch, NotMatch_obj, NotMatch_sheet, NotMatch_rows = open_file('NotMatch')
ZPartNumber_nm, ZCompanyName_nm, ZReplacementPart_nm, ZReplacementCompany_nm, OnlineSource_nm, ZURLSource_nm, comment_nm = get_notmatch_data(f2_indecies, work_sheet, headers)
make_notmatch_file(NotMatch_sheet, ZPartNumber_nm, ZCompanyName_nm, ZReplacementPart_nm, ZReplacementCompany_nm, OnlineSource_nm, ZURLSource_nm, comment_nm)

# obs file
obs, obs_obj, obs_sheet, obs_rows = open_file('obs')
ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs = get_obs_data(work_sheet, work_rows, headers)
make_obs_file(ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs)


# part partsRepInsert
partsRepInsert, partsRepInsert_obj, partsRepInsert_sheet, partsRepInsert_rows = open_file('PartsReplacementInsert')
ZPartNumber_obs , ZCompanyName_obs, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs, SourceType = get_partsRepInsert_data(work_sheet, work_rows,  headers)
NewReplacementType, ReplacementFeature = get_NewReplaceType(ObsolescenceCode_obs)
ZReplacementPart, ZReplacementCompany, DiffFeatures, PinToPin = handle_exportedfilesANDpartsRepInsert_file(ObsolescenceCode_obs, ZPartNumber_obs, HasReplacement_obs, ZCompanyName_obs)
# Exceptionflag = get_Exceptionflag(ObsolescenceCode_obs, OffLineSource1_obs)
make_partsRepInsert_file(partsRepInsert_sheet, ZPartNumber_obs , ZCompanyName_obs, ZReplacementPart, ZReplacementCompany, LCStatus_obs, ObsolescenceCode_obs, ObsolescenceReason_obs, DiffFeatures, HasReplacement_obs, Critical_obs, OnLineSource1_obs, OffLineSource1_obs, SourceType_obs, SourceType, NewReplacementType, ReplacementFeature,  PinToPin, exportedMask=True)














